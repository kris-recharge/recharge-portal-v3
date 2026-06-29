"""GET /api/export — XLSX download of charging sessions + vendor faults.

v3.2: Timestamps are now written as native Excel datetime cells (not strings) so
Excel recognises them without manual reformatting; Vendor Error Code is written as
a real number (no "Number stored as Text" warning); rows are ordered by session
START time, newest on top; and the CSV format option has been removed (xlsx only).
"""

from __future__ import annotations

import math
import io
from datetime import datetime, timezone
from zoneinfo import ZoneInfo

from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse

from ..auth import CurrentUser, filter_evse_ids
from ..config import DEV_BYPASS_AUTH
from ..constants import connector_type_for, display_name, get_all_station_ids, location_label
from ..db import acquire
from .utility import ADMIN_EMAIL, UTILITY_EFFICIENCY_SQL  # v3.2: shared metered-vs-dispensed query

router = APIRouter(prefix="/api/export", tags=["export"])

_AK = ZoneInfo("America/Anchorage")

# v3.2: Excel display format for the datetime columns — "m/dd/yy h:mm:ss".
_XLSX_DT_FORMAT = "m/dd/yy h:mm:ss"


def _ak_dt(dt: datetime | None) -> datetime | None:
    """Convert a UTC datetime to a naive Alaska-local datetime for Excel.

    v3.2: returns a real datetime (Excel stores it as a date serial) instead of a
    preformatted string, so Excel recognises the cell as a date/time natively.
    """
    if dt is None:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    # Drop tzinfo — Excel has no concept of timezone; values are AK-local wall time.
    return dt.astimezone(_AK).replace(tzinfo=None)


def _pct(val) -> str:
    """Format SoC value (0–100) as '45%' string; blank if absent."""
    if val is None or val == "":
        return ""
    return f"{float(val):.0f}%"


_CC_TAGS = {"FE6DD7B2C3904F", "161D77C442099C"}


def _auth_method(auth_tag: str) -> str:
    """Derive Authentication Method from the raw auth tag.

    VID:*  → AutoCharge (vehicle-initiated via Autocharge protocol)
    known CC tag IDs → CC
    anything else    → App
    """
    if not auth_tag:
        return "App"
    if auth_tag.startswith("VID:"):
        return "AutoCharge"
    if auth_tag in _CC_TAGS:
        return "CC"
    return "App"


@router.get("")
async def export_sessions(
    user: CurrentUser,
    start_date: str = Query(..., description="YYYY-MM-DD Alaska local"),
    end_date: str = Query(..., description="YYYY-MM-DD Alaska local"),
    station_id: list[str] | None = Query(None),
    format: str = Query("xlsx", pattern="^(xlsx)$"),  # v3.2: CSV removed, xlsx only
):
    all_ids = get_all_station_ids()
    allowed = filter_evse_ids(all_ids, user.allowed_evse_ids)
    if station_id:
        allowed = [s for s in station_id if s in allowed]

    start_utc = (
        datetime.fromisoformat(f"{start_date}T00:00:00")
        .replace(tzinfo=_AK)
        .astimezone(timezone.utc)
    )
    end_utc = (
        datetime.fromisoformat(f"{end_date}T23:59:59")
        .replace(tzinfo=_AK)
        .astimezone(timezone.utc)
    )

    async with acquire() as conn:
        # ── Sessions ──────────────────────────────────────────────────────────
        rows = await conn.fetch(
            """
            WITH sessions AS (
                SELECT
                    m.station_id,
                    m.connector_id,
                    m.transaction_id,
                    MIN(m.ts)                             AS start_utc,
                    MAX(m.ts)                             AS end_utc,
                    MAX(m.power_w)                        AS max_power_w,
                    MAX(m.energy_wh) - MIN(m.energy_wh)  AS energy_wh_delta,
                    -- first SoC reading
                    (SELECT mv2.soc FROM meter_values_parsed mv2
                     WHERE mv2.station_id    = m.station_id
                       AND mv2.connector_id  = m.connector_id
                       AND mv2.transaction_id = m.transaction_id
                       AND mv2.soc IS NOT NULL
                     ORDER BY mv2.ts ASC LIMIT 1)           AS soc_start,
                    -- first non-zero SoC reading (used to skip bogus leading 0%)
                    (SELECT mv2.soc FROM meter_values_parsed mv2
                     WHERE mv2.station_id    = m.station_id
                       AND mv2.connector_id  = m.connector_id
                       AND mv2.transaction_id = m.transaction_id
                       AND mv2.soc IS NOT NULL
                       AND mv2.soc > 0
                     ORDER BY mv2.ts ASC LIMIT 1)           AS soc_first_nonzero,
                    -- last SoC reading
                    (SELECT mv2.soc FROM meter_values_parsed mv2
                     WHERE mv2.station_id    = m.station_id
                       AND mv2.connector_id  = m.connector_id
                       AND mv2.transaction_id = m.transaction_id
                       AND mv2.soc IS NOT NULL
                     ORDER BY mv2.ts DESC LIMIT 1)          AS soc_end
                FROM meter_values_parsed m
                WHERE m.station_id     = ANY($1::text[])
                  AND m.transaction_id IS NOT NULL
                  AND m.ts >= $2
                  -- v3.2: filter by session START time. Widen the reading window
                  -- 1 day past end so sessions that START in range are aggregated
                  -- in full (not truncated); the outer WHERE drops later starts.
                  AND m.ts <= $3::timestamptz + INTERVAL '1 day'
                GROUP BY m.station_id, m.connector_id, m.transaction_id
            )
            with_auth AS (
                SELECT
                    s.*,
                    -- Authentication: raw id_tag from authorize_methods (all methods)
                    (SELECT a.id_tag FROM authorize_methods a
                     WHERE a.asset_id = s.station_id
                       AND a.start_received_at BETWEEN s.start_utc - INTERVAL '60 minutes'
                                                   AND s.start_utc + INTERVAL '5 minutes'
                     ORDER BY ABS(EXTRACT(EPOCH FROM (a.start_received_at - s.start_utc))) ASC
                     LIMIT 1) AS auth_tag,
                    -- VID: only VID:-prefixed tags from ocpp_events Authorize
                    (SELECT oe.action_payload->>'idTag'
                     FROM ocpp_events oe
                     WHERE oe.asset_id = s.station_id
                       AND oe.action   = 'Authorize'
                       AND oe.action_payload->>'idTag' LIKE 'VID:%'
                       AND oe.received_at BETWEEN s.start_utc - INTERVAL '60 minutes'
                                              AND s.start_utc + INTERVAL '5 minutes'
                     ORDER BY ABS(EXTRACT(EPOCH FROM (oe.received_at - s.start_utc))) ASC
                     LIMIT 1) AS vid_tag,
                    ep.price_per_kwh,
                    ep.connection_fee
                FROM sessions s
                LEFT JOIN LATERAL (
                    SELECT price_per_kwh, connection_fee
                    FROM evse_pricing
                    WHERE station_id = s.station_id
                      AND effective_start <= s.start_utc
                      AND (effective_end IS NULL OR effective_end > s.start_utc)
                    ORDER BY effective_start DESC LIMIT 1
                ) ep ON true
            ),
            real_rows AS (
                SELECT
                    'session'::text                     AS kind,
                    station_id,
                    connector_id,
                    transaction_id::text                AS transaction_id,
                    start_utc,
                    end_utc,
                    max_power_w::numeric                AS max_power_w,
                    energy_wh_delta::numeric            AS energy_wh_delta,
                    soc_start::numeric                  AS soc_start,
                    soc_first_nonzero::numeric          AS soc_first_nonzero,
                    soc_end::numeric                    AS soc_end,
                    auth_tag,
                    vid_tag,
                    price_per_kwh::numeric              AS price_per_kwh,
                    connection_fee::numeric            AS connection_fee
                FROM with_auth
                WHERE start_utc <= $3   -- v3.2: keep only sessions that START in range
            ),
            -- ── Failed start attempts (mirrors /api/sessions) ───────────────────
            -- A plug-in (StatusNotification "Preparing") where the driver presented
            -- a real credential (non-VID token Authorize via CC reader or app) but
            -- the session never reached "Charging" before the connector cleared
            -- (capped at 2 h). These never mint a transaction_id, so they never land
            -- in meter_values_parsed and are otherwise invisible in the export.
            -- Excluded by design: AutoCharge VID:* probes and plug-and-unplug blips
            -- with no Authorize at all.
            sn AS (
                SELECT
                    asset_id,
                    connector_id,
                    received_at,
                    action_payload->>'status'                        AS status,
                    LAG(action_payload->>'status') OVER (
                        PARTITION BY asset_id, connector_id ORDER BY received_at
                    )                                                AS prev_status
                FROM ocpp_events
                WHERE asset_id = ANY($1::text[])
                  AND action = 'StatusNotification'
                  AND action_payload->>'status' IS NOT NULL
            ),
            charge_sig AS (   -- any signal that a transaction actually began
                SELECT asset_id, connector_id, received_at
                FROM ocpp_events
                WHERE asset_id = ANY($1::text[])
                  AND ( (action = 'StatusNotification'
                         AND action_payload->>'status' = 'Charging')
                        OR action = 'StartTransaction' )
            ),
            avail AS (        -- connector cleared / unplugged
                SELECT asset_id, connector_id, received_at
                FROM ocpp_events
                WHERE asset_id = ANY($1::text[])
                  AND action = 'StatusNotification'
                  AND action_payload->>'status' = 'Available'
            ),
            attempts AS (
                SELECT
                    s.asset_id    AS station_id,
                    s.connector_id,
                    s.received_at AS attempt_at,
                    LEAST(
                        s.received_at + INTERVAL '2 hours',
                        COALESCE((SELECT MIN(a.received_at) FROM avail a
                                  WHERE a.asset_id = s.asset_id
                                    AND a.connector_id IS NOT DISTINCT FROM s.connector_id
                                    AND a.received_at > s.received_at),
                                 'infinity'::timestamptz)
                    )                                            AS episode_end
                FROM sn s
                WHERE s.status = 'Preparing'
                  AND s.prev_status IS DISTINCT FROM 'Preparing'   -- first Preparing of an episode
                  AND s.received_at >= $2
                  AND s.received_at <= $3
            ),
            attempts_filtered AS (
                SELECT a.station_id, a.connector_id, a.attempt_at
                FROM attempts a
                WHERE
                    NOT EXISTS (   -- (1) never reached Charging within the episode
                        SELECT 1 FROM charge_sig c
                        WHERE c.asset_id = a.station_id
                          AND c.connector_id IS NOT DISTINCT FROM a.connector_id
                          AND c.received_at > a.attempt_at
                          AND c.received_at < a.episode_end
                    )
                    AND EXISTS (   -- (2) but a real, non-AutoCharge credential was presented
                        SELECT 1 FROM ocpp_events az
                        WHERE az.asset_id = a.station_id
                          AND az.action = 'Authorize'
                          AND az.action_payload->>'idTag' NOT LIKE 'VID:%'
                          AND az.received_at BETWEEN a.attempt_at - INTERVAL '30 seconds'
                                                 AND a.episode_end
                    )
            ),
            failed_rows AS (
                SELECT
                    'failed'::text                      AS kind,
                    station_id,
                    connector_id,
                    'attempt:' || station_id || ':' || COALESCE(connector_id, 0)
                               || ':' || EXTRACT(EPOCH FROM attempt_at)::bigint::text AS transaction_id,
                    attempt_at                          AS start_utc,
                    NULL::timestamptz                   AS end_utc,
                    NULL::numeric                       AS max_power_w,
                    NULL::numeric                       AS energy_wh_delta,
                    NULL::numeric                       AS soc_start,
                    NULL::numeric                       AS soc_first_nonzero,
                    NULL::numeric                       AS soc_end,
                    NULL::text                          AS auth_tag,
                    NULL::text                          AS vid_tag,
                    NULL::numeric                       AS price_per_kwh,
                    NULL::numeric                       AS connection_fee
                FROM attempts_filtered
            ),
            unioned AS (
                SELECT * FROM real_rows
                UNION ALL
                SELECT * FROM failed_rows
            )
            SELECT * FROM unioned
            ORDER BY start_utc DESC NULLS LAST  -- v3.2: newest start on top
            """,
            allowed, start_utc, end_utc,
        )

        # ── Vendor Faults ─────────────────────────────────────────────────────
        fault_rows = await conn.fetch(
            """
            SELECT
                oe.received_at,
                oe.asset_id,
                oe.connector_id,
                oe.action_payload->>'errorCode'       AS error_code,
                oe.action_payload->>'vendorErrorCode' AS vendor_error_code,
                oe.action_payload->>'status'          AS status,
                -- Vendor description scoped to manufacturer: Tritium units use
                -- tritium_error_codes, Alpitronic (HYC400) use alpitronic_error_codes.
                CASE
                    WHEN ut.manufacturer = 'Tritium'    THEN tec.description
                    WHEN ut.manufacturer = 'Alpitronic' THEN aec.description
                    ELSE NULL
                END                                   AS vendor_description
            FROM ocpp_events oe
            LEFT JOIN chargers c    ON c.external_id = oe.asset_id
            LEFT JOIN unit_types ut ON ut.id = c.unit_type_id
            LEFT JOIN tritium_error_codes tec
                ON ut.manufacturer = 'Tritium'
               AND tec.code = (
                    CASE WHEN oe.action_payload->>'vendorErrorCode' ~ '^[0-9]+$'
                         THEN (oe.action_payload->>'vendorErrorCode')::integer
                    END
                )
            LEFT JOIN alpitronic_error_codes aec
                ON ut.manufacturer = 'Alpitronic'
               AND aec.error_code = (
                    CASE WHEN oe.action_payload->>'vendorErrorCode' ~ '^[0-9]+$'
                         THEN (oe.action_payload->>'vendorErrorCode')::integer
                    END
                )
            WHERE oe.asset_id = ANY($1::text[])
              AND oe.action   = 'StatusNotification'
              AND oe.action_payload->>'errorCode' != 'NoError'
              AND oe.received_at >= $2
              AND oe.received_at <= $3
            ORDER BY oe.received_at DESC
            """,
            allowed, start_utc, end_utc,
        )

        # ── Utility Reads (v3.2) — daily metered vs dispensed per meter ────────
        # Charger-aware: a meter narrows to one unit (charger_id) or covers a
        # whole site. Newest day on top, scoped to the user's allowed EVSEs.
        # Unrestricted users exporting without an EVSE filter also get
        # metered-only meters (no site mapping, e.g. CEA 2630156). Mirror the
        # /efficiency endpoint: admin email and DEV_BYPASS count as unrestricted
        # even when the user has an explicit allowed_evse_ids list.
        is_admin = user.email == ADMIN_EMAIL or DEV_BYPASS_AUTH
        unrestricted_export = (
            (is_admin
             or user.allowed_evse_ids is None
             or "__ALL__" in (user.allowed_evse_ids or []))
            and not station_id
        )
        utility_rows = await conn.fetch(
            UTILITY_EFFICIENCY_SQL.replace(
                "ORDER BY m.site_name, m.account_number, u.day",
                "ORDER BY m.site_name, m.account_number, u.day DESC",
            ),
            unrestricted_export,  # $1
            allowed,              # $2 allowed station ids
            start_utc,            # $3 (timestamptz cast to ::date)
            end_utc,              # $4
        )

    # ── Build sessions rows ───────────────────────────────────────────────────
    session_columns = [
        "Status",                 # A — Completed | Auth Timed Out (v3.2)
        "Start Date/Time (AK)",   # B
        "End Date/Time (AK)",     # C
        "EVSE",                   # D
        "Location",               # E
        "Connector #",            # F
        "Connector Type",         # G
        "Max Power (kW)",         # H
        "Energy Delivered (kWh)", # I
        "Duration (min)",         # J
        "SoC Start (%)",          # K
        "SoC End (%)",            # L
        "Authentication",         # M
        "Authentication Method",  # N
        "Est. Revenue (USD)",     # O
        "VID",                    # P
    ]

    data_rows: list[list] = []
    for r in rows:
        sid        = r["station_id"]
        conn_id    = r["connector_id"]
        start_dt   = r["start_utc"]
        end_dt     = r["end_utc"]
        is_failed  = r["kind"] == "failed"
        # Failed-start attempts carry no transaction → no energy/SoC/revenue.
        # Match the on-screen badge so export counts reconcile with the table.
        status     = "Auth Timed Out" if is_failed else "Completed"
        dur_min    = (
            round((end_dt - start_dt).total_seconds() / 60.0, 2)
            if start_dt and end_dt else ""
        )
        energy_wh  = r["energy_wh_delta"]
        energy_kwh = round(float(energy_wh) / 1000.0, 3) if energy_wh else ""
        max_kw     = round(float(r["max_power_w"]) / 1000.0, 2) if r["max_power_w"] else ""
        p_kwh      = float(r["price_per_kwh"] or 0)
        c_fee      = float(r["connection_fee"] or 0)
        est_rev    = (
            math.floor((c_fee + (float(energy_wh or 0) / 1000.0) * p_kwh) * 100) / 100
            if (p_kwh or c_fee) else ""
        )

        # SoC — mirror _resolve_soc() from sessions.py:
        # 1. Scale normalisation: chargers that report 0-1 fraction instead of 0-100
        soc_start_raw     = r["soc_start"]
        soc_first_nonzero = r["soc_first_nonzero"]
        soc_end_raw       = r["soc_end"]
        ref = soc_end_raw if soc_end_raw is not None else soc_start_raw
        scale = 100.0 if (ref is not None and float(ref) <= 1.0) else 1.0

        soc_end_val = round(float(soc_end_raw) * scale, 1) if soc_end_raw is not None else None

        # 2. Bogus-zero filter: skip leading 0% readings if first non-zero > 1.5%
        soc_start_val: float | None = None
        if soc_start_raw is not None:
            soc_start_val = float(soc_start_raw) * scale
            if soc_start_val == 0.0 and soc_first_nonzero is not None:
                nonzero_val = float(soc_first_nonzero) * scale
                if nonzero_val > 1.5:
                    soc_start_val = nonzero_val
            soc_start_val = round(soc_start_val, 1)

        data_rows.append([
            status,             # A — Status (v3.2)
            _ak_dt(start_dt),   # B — native datetime cell (v3.2)
            _ak_dt(end_dt),     # C — native datetime cell (v3.2)
            display_name(sid),
            location_label(sid),
            conn_id,
            connector_type_for(sid, conn_id or 0, start_dt),
            max_kw,
            energy_kwh,
            dur_min,
            _pct(soc_start_val),
            _pct(soc_end_val),
            r["auth_tag"] or "",                        # M — Authentication (raw tag)
            _auth_method(r["auth_tag"] or ""),          # N — Authentication Method
            est_rev,                                    # O — Est. Revenue
            r["vid_tag"] or "",                         # P — VID (VID:-prefixed only)
        ])

    # ── Build faults rows ─────────────────────────────────────────────────────
    fault_columns = [
        "Timestamp (AK)",
        "EVSE",
        "Location",
        "Connector #",
        "Error Code",
        "Vendor Error Code",
        "Description",
        "Status",
    ]
    fault_data: list[list] = []
    for fr in fault_rows:
        # v3.2: Vendor Error Code (col F) written as a real number when numeric,
        # so Excel stops flagging "Number stored as Text".
        vec_raw = fr["vendor_error_code"]
        vec: int | str = int(vec_raw) if (vec_raw and str(vec_raw).isdigit()) else (vec_raw or "")
        fault_data.append([
            _ak_dt(fr["received_at"]),   # A — native datetime cell (v3.2)
            display_name(fr["asset_id"]),
            location_label(fr["asset_id"]),
            fr["connector_id"] or "",
            fr["error_code"] or "",
            vec,                          # F — numeric Vendor Error Code (v3.2)
            fr["vendor_description"] or "",
            fr["status"] or "",
        ])

    # ── Build Utility Reads rows (v3.2) ───────────────────────────────────────
    utility_columns = [
        "Date",                 # A
        "Site",                 # B
        "Unit",                 # C (blank = whole-site meter)
        "Utility",              # D
        "Account #",            # E
        "Metered kWh",          # F
        "Dispensed kWh",        # G
        "Efficiency (%)",       # H
    ]
    utility_data: list[list] = []
    for ur in utility_rows:
        metered   = float(ur["metered_kwh"]) if ur["metered_kwh"] is not None else 0.0
        # Metered-only meters (no site mapping) have no dispensed side: blank cells
        dispensed = float(ur["dispensed_kwh"]) if ur["dispensed_kwh"] is not None else None
        eff = (
            round(dispensed / metered * 100.0, 1)
            if dispensed is not None and metered > 0 else ""
        )
        utility_data.append([
            ur["day"],                                     # A — native date cell
            ur["site_name"] or ur["display_name"] or "",   # site-less: display name
            ur["unit_name"] or "",
            ur["utility"],
            ur["account_number"],
            round(metered, 3),
            round(dispensed, 3) if dispensed is not None else "",
            eff,
        ])

    filename = f"sessions_{start_date}_to_{end_date}"

    # ── XLSX: two sheets (v3.2: CSV option removed) ───────────────────────────
    import openpyxl

    wb = openpyxl.Workbook()

    # Sheet 1 — Sessions
    ws1 = wb.active
    ws1.title = "Sessions"
    ws1.append(session_columns)
    for row in data_rows:
        ws1.append(row)
    # v3.2: format the two datetime columns (B=Start, C=End) as real dates.
    for col in ("B", "C"):
        for cell in ws1[col][1:]:   # skip header row
            cell.number_format = _XLSX_DT_FORMAT

    # Sheet 2 — Vendor Faults
    ws2 = wb.create_sheet(title="Vendor Faults")
    ws2.append(fault_columns)
    for row in fault_data:
        ws2.append(row)
    # v3.2: format the timestamp column (A) as a real date.
    for cell in ws2["A"][1:]:
        cell.number_format = _XLSX_DT_FORMAT

    # Sheet 3 — Utility Reads (v3.2)
    ws3 = wb.create_sheet(title="Utility Reads")
    ws3.append(utility_columns)
    for row in utility_data:
        ws3.append(row)
    # Date column (A) as a real date.
    for cell in ws3["A"][1:]:
        cell.number_format = "m/dd/yy"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


# ── Maintenance Activities export (v3.2) ──────────────────────────────────────


def _task_response(r) -> str:
    """Render a single PM task answer into one human-readable cell.

    Tasks vary by input_type: pass/fail, checkbox (completed), measurement, or
    free text. Pick whichever field the tech actually filled in.
    """
    if r["result_pass_fail"]:
        return str(r["result_pass_fail"])          # pass | fail | na
    if r["result_completed"] is not None:
        return "Completed" if r["result_completed"] else "Not completed"
    mv = r["result_measured_value"]
    if mv not in (None, ""):
        return str(mv)
    if r["result_text"]:
        return str(r["result_text"])
    return ""


@router.get("/maintenance")
async def export_maintenance(
    user: CurrentUser,
    start_date: str = Query(..., description="YYYY-MM-DD Alaska local"),
    end_date: str = Query(..., description="YYYY-MM-DD Alaska local"),
    station_id: list[str] | None = Query(None),
    format: str = Query("xlsx", pattern="^(xlsx)$"),  # xlsx only, mirrors sessions export
):
    """XLSX dump of PM/maintenance logs (header, every Q&A, parts) for warranty proof.

    EVSE scoping mirrors the sessions export exactly so a user can never reach PM
    logs for a unit they are not approved for in portal_users.allowed_evse_ids:

      • Non-admin → records are scoped to the intersection of their approved EVSEs
        and any chip selection (WHERE c.external_id = ANY(allowed)). Units with a
        NULL external_id (non-LynkWell, e.g. Terra184) can never match.
      • Admin (ADMIN_EMAIL / DEV_BYPASS) with no chip filter → no external_id
        restriction, so NULL-external_id units are included. When an admin DOES
        select chips, the same allowed-list filter applies.
    """
    is_admin = user.email == ADMIN_EMAIL or DEV_BYPASS_AUTH

    all_ids = get_all_station_ids()
    allowed = filter_evse_ids(all_ids, user.allowed_evse_ids)
    if station_id:
        allowed = [s for s in station_id if s in allowed]

    start_utc = (
        datetime.fromisoformat(f"{start_date}T00:00:00")
        .replace(tzinfo=_AK)
        .astimezone(timezone.utc)
    )
    end_utc = (
        datetime.fromisoformat(f"{end_date}T23:59:59")
        .replace(tzinfo=_AK)
        .astimezone(timezone.utc)
    )

    # Admins dumping everything (no chip filter) skip the external_id allowlist so
    # NULL-external_id units appear; everyone else is restricted to `allowed`.
    unrestricted = is_admin and not station_id

    async with acquire() as conn:
        if unrestricted:
            record_rows = await conn.fetch(
                """
                SELECT
                    mr.id::text AS id,
                    mr.record_timestamp,
                    mr.record_type,
                    mr.technician_name,
                    mr.onsite_hours,
                    mr.mobilized_hours,
                    mr.overall_result,
                    mr.firmware_version,
                    mr.work_description,
                    mr.additional_work_needed,
                    mr.planned_future_work,
                    c.external_id,
                    c.name AS charger_name,
                    s.name AS site_name
                FROM maintenance_records mr
                JOIN chargers c ON c.id = mr.charger_id
                LEFT JOIN sites s ON s.id = mr.site_id
                WHERE mr.record_timestamp >= $1
                  AND mr.record_timestamp <= $2
                ORDER BY mr.record_timestamp DESC
                """,
                start_utc, end_utc,
            )
        else:
            record_rows = await conn.fetch(
                """
                SELECT
                    mr.id::text AS id,
                    mr.record_timestamp,
                    mr.record_type,
                    mr.technician_name,
                    mr.onsite_hours,
                    mr.mobilized_hours,
                    mr.overall_result,
                    mr.firmware_version,
                    mr.work_description,
                    mr.additional_work_needed,
                    mr.planned_future_work,
                    c.external_id,
                    c.name AS charger_name,
                    s.name AS site_name
                FROM maintenance_records mr
                JOIN chargers c ON c.id = mr.charger_id
                LEFT JOIN sites s ON s.id = mr.site_id
                WHERE mr.record_timestamp >= $1
                  AND mr.record_timestamp <= $2
                  AND c.external_id = ANY($3::text[])
                ORDER BY mr.record_timestamp DESC
                """,
                start_utc, end_utc, allowed,
            )

        record_ids = [r["id"] for r in record_rows]

        # Task results + the question text they answer. Scoped to the records we
        # already authorised above, so there's no back-door to other units' Q&A.
        task_rows = []
        part_rows = []
        if record_ids:
            task_rows = await conn.fetch(
                """
                SELECT
                    tr.record_id::text AS record_id,
                    t.task_category,
                    t.task_order,
                    t.task_name,
                    t.unit_of_measure,
                    t.critical_fail,
                    tr.result_pass_fail,
                    tr.result_completed,
                    tr.result_measured_value,
                    tr.result_text,
                    tr.task_notes
                FROM pm_task_results tr
                LEFT JOIN pm_template_tasks t ON t.id = tr.task_id
                WHERE tr.record_id = ANY($1::uuid[])
                ORDER BY tr.record_id, t.task_order
                """,
                record_ids,
            )
            part_rows = await conn.fetch(
                """
                SELECT
                    record_id::text AS record_id,
                    part_name,
                    part_number,
                    action_taken,
                    notes
                FROM maintenance_parts_replaced
                WHERE record_id = ANY($1::uuid[])
                ORDER BY record_id, part_name
                """,
                record_ids,
            )

    def _evse_label(external_id, charger_name) -> str:
        return display_name(external_id) if external_id else (charger_name or "")

    def _loc_label(external_id, site_name) -> str:
        return location_label(external_id) if external_id else (site_name or "")

    # Submitted-timestamp lookup per record, reused on the detail sheets.
    submitted_by_id = {r["id"]: _ak_dt(r["record_timestamp"]) for r in record_rows}
    evse_by_id = {r["id"]: _evse_label(r["external_id"], r["charger_name"]) for r in record_rows}

    # ── Sheet 1: PM Logs (one row per log) ────────────────────────────────────
    log_columns = [
        "Log ID", "EVSE", "Location", "Record Type", "Technician",
        "Submitted (AK)", "Onsite Hrs", "Mobilized Hrs", "Overall Result",
        "Firmware", "Work Description", "Additional Work Needed",
        "Planned Future Work",
    ]
    log_data: list[list] = []
    for r in record_rows:
        log_data.append([
            r["id"],
            _evse_label(r["external_id"], r["charger_name"]),
            _loc_label(r["external_id"], r["site_name"]),
            r["record_type"] or "",
            r["technician_name"] or "",
            _ak_dt(r["record_timestamp"]),
            float(r["onsite_hours"]) if r["onsite_hours"] is not None else "",
            float(r["mobilized_hours"]) if r["mobilized_hours"] is not None else "",
            r["overall_result"] or "",
            r["firmware_version"] or "",
            r["work_description"] or "",
            "Yes" if r["additional_work_needed"] else "No",
            r["planned_future_work"] or "",
        ])

    # ── Sheet 2: Task Results (one row per question) ──────────────────────────
    task_columns = [
        "Log ID", "EVSE", "Submitted (AK)", "Category", "Question",
        "Response", "Measured Value", "Unit", "Critical", "Notes",
    ]
    task_data: list[list] = []
    for tr in task_rows:
        rid = tr["record_id"]
        mv = tr["result_measured_value"]
        task_data.append([
            rid,
            evse_by_id.get(rid, ""),
            submitted_by_id.get(rid),
            tr["task_category"] or "",
            tr["task_name"] or "",
            _task_response(tr),
            str(mv) if mv not in (None, "") else "",
            tr["unit_of_measure"] or "",
            "Yes" if tr["critical_fail"] else "",
            tr["task_notes"] or "",
        ])

    # ── Sheet 3: Parts Replaced (one row per part) ────────────────────────────
    part_columns = [
        "Log ID", "EVSE", "Submitted (AK)", "Part Name", "Part Number",
        "Action Taken", "Notes",
    ]
    part_data: list[list] = []
    for p in part_rows:
        rid = p["record_id"]
        part_data.append([
            rid,
            evse_by_id.get(rid, ""),
            submitted_by_id.get(rid),
            p["part_name"] or "",
            p["part_number"] or "",
            p["action_taken"] or "",
            p["notes"] or "",
        ])

    import openpyxl

    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "PM Logs"
    ws1.append(log_columns)
    for row in log_data:
        ws1.append(row)
    for cell in ws1["F"][1:]:   # Submitted (AK)
        cell.number_format = _XLSX_DT_FORMAT

    ws2 = wb.create_sheet(title="Task Results")
    ws2.append(task_columns)
    for row in task_data:
        ws2.append(row)
    for cell in ws2["C"][1:]:   # Submitted (AK)
        cell.number_format = _XLSX_DT_FORMAT

    ws3 = wb.create_sheet(title="Parts Replaced")
    ws3.append(part_columns)
    for row in part_data:
        ws3.append(row)
    for cell in ws3["C"][1:]:   # Submitted (AK)
        cell.number_format = _XLSX_DT_FORMAT

    filename = f"maintenance_{start_date}_to_{end_date}"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )
